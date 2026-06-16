"""
Export discovered devices to CSV, Excel (.xlsx) and JSON.

Each exporter accepts a list of device dicts (as returned by
``DiscoveryStore.list_devices`` or ``ScanResult.hosts``) and writes a file,
returning the path written.
"""
from __future__ import annotations

import csv
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

_COLUMNS = [
    ("ip_address", "IP Address"),
    ("mac_address", "MAC Address"),
    ("hostname", "Hostname"),
    ("vendor", "Vendor"),
    ("device_type", "Device Type"),
    ("open_ports", "Open Ports"),
    ("status", "Status"),
    ("first_seen", "First Seen"),
    ("last_seen", "Last Seen"),
]


def _timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def _row(device: dict[str, Any]) -> dict[str, Any]:
    row = {}
    for key, _label in _COLUMNS:
        val = device.get(key)
        if key == "open_ports" and isinstance(val, (list, tuple)):
            val = ", ".join(str(p) for p in val)
        row[key] = "" if val is None else val
    return row


def export_csv(devices: Iterable[dict[str, Any]], out_dir: Path, name: str | None = None) -> Path:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / (name or f"devices_{_timestamp()}.csv")
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[k for k, _ in _COLUMNS])
        writer.writerow({k: label for k, label in _COLUMNS})  # header labels
        for d in devices:
            writer.writerow(_row(d))
    return path


def export_json(devices: Iterable[dict[str, Any]], out_dir: Path, name: str | None = None) -> Path:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / (name or f"devices_{_timestamp()}.json")
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "count": 0,
        "devices": [],
    }
    devices = list(devices)
    payload["count"] = len(devices)
    payload["devices"] = devices
    path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    return path


def export_excel(devices: Iterable[dict[str, Any]], out_dir: Path, name: str | None = None) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for Excel export (pip install openpyxl)") from exc

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / (name or f"devices_{_timestamp()}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Network Devices"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    online_fill = PatternFill("solid", fgColor="DCFCE7")
    offline_fill = PatternFill("solid", fgColor="FEE2E2")

    labels = [label for _k, label in _COLUMNS]
    ws.append(labels)
    for col, _ in enumerate(labels, start=1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    devices = list(devices)
    for d in devices:
        r = _row(d)
        ws.append([r[k] for k, _ in _COLUMNS])
        status = (d.get("status") or "").lower()
        fill = online_fill if status == "online" else (offline_fill if status == "offline" else None)
        if fill:
            for col in range(1, len(_COLUMNS) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill

    # Auto-size columns (approximate)
    for col, (key, label) in enumerate(_COLUMNS, start=1):
        width = max(len(label), 12)
        for d in devices:
            width = max(width, len(str(_row(d)[key])))
        ws.column_dimensions[get_column_letter(col)].width = min(width + 2, 45)

    ws.freeze_panes = "A2"
    wb.save(path)
    return path


def export_all(devices: Iterable[dict[str, Any]], out_dir: Path) -> dict[str, str]:
    devices = list(devices)
    return {
        "csv": str(export_csv(devices, out_dir)),
        "json": str(export_json(devices, out_dir)),
        "excel": str(export_excel(devices, out_dir)),
    }
